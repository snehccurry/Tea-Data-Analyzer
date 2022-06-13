from tkinter import *
from PIL import ImageTk, Image
import numpy as np
import time
import pandas as pd
import os
from datetime import datetime
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib import pyplot as plt
from matplotlib import style
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
from tkinter import ttk, filedialog
import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook







initialdir="C:\\Users\\"+os.getlogin()+"\\Documents"
#print(initialdir)

###################create a workbook instance
wb=Workbook()


####################load exsisting workbook
wb=load_workbook('officedata.xlsx')



#####################create active worksheet
ws=wb.active


column_a=ws['A']
column_a=ws['B']



root=Tk()
root.title('Factory Analysis')
#root.iconbitmap('path')
root.geometry("400x200")




#################################################################################						Main Frame begins
main_frame=LabelFrame(root, bd=0, bg="#000000")
label1=Label(main_frame,text="Burtoll Tea Analysis", bd=0).pack()
main_frame.pack()

################################################################################						Main Frame ends








################################################# excel frame 

my_frame= Frame(root, bg="#E2A089")
my_frame.pack(pady=20)


my_tree=ttk.Treeview(my_frame,)
my_tree.config(height=40)

#file open function

def file_open():
	filename=filedialog.askopenfilename(initialdir=initialdir, title="Open Excel file", filetype=(("xlsx files","*.xlsx"),("All Files","*.*")))
	if filename:
		try:
			filename=r"{}".format(filename)
			#df= pd.read_excel(filename) #for release
		except ValueError:
			error_label.config(text="An error occured")
		except FileNotFoundError:
			error_label.config(text="File not found")
	#clear old treeview
	clear_tree()
	create_tree_view_for_excel_sheet()
	

def create_tree_view_for_excel_sheet():
	#set up new treeview
	my_tree["column"]=list(df.columns)
	my_tree["show"]="headings"
	#loop through column list for headers

	for column in my_tree["column"]:
		my_tree.heading(column,text=column)

	#data in treeview

	df_rows=df.to_numpy().tolist()
	for row in df_rows:
		my_tree.insert("","end",values=row)
	my_tree.pack()








def clear_tree():
	my_tree.delete(*my_tree.get_children())

#add menu

my_menu=Menu(root)
root.config(menu=my_menu)

# Add menu Dropdown
file_menu= Menu(my_menu, tearoff=False)
my_menu.add_cascade(label="Spreadsheets", menu=file_menu)
file_menu.add_command(label="Open", command=file_open)

################################################# excel frame ends here














##################################################################################							weekly Graph
def Weekly_Graph():
	'''x=[1,2,3]
				y=[1,4,9]
				plt.plot(x,y)
				plt.show()'''
	# Read CSV into pandas
	#data = pd.read_csv(r"officedata.csv")
	data=pd.read_excel(r"officedata.xlsx")
	data.head()
	df = pd.DataFrame(data)
	##print(df)
	#df['ConvertedDate']=df['Date'].astype(str)
	Date= df['Date'].dt.strftime('%d-%b-%Y').head(7)

	#print(Date)
	TeaPlucked = df['Tea Plucked'].head(7)
	#print(TeaPlucked)
	TeaMade=df['Tea Made'].head(7)
	#print(TeaMade)
	# Figure Size
	fig = plt.figure(figsize =(7, 7))
	# Horizontal Bar Plot
	ypos=np.arange(len(Date))
	#print("length of ypos is: ",ypos)
	for i in range(len(ypos)):
		plt.text(i,TeaMade[i],TeaMade[i])
	#plt.bar(ypos, TeaMade, color='#9B23D0',width=0.5)
	plt.bar(ypos, TeaMade, color='#A934BD',width=0.5)
	#plt.bar(ypos, TeaMade, color='#C6808B',width=0.5)
	
	plt.plot(ypos, TeaPlucked, color='#EC745C')
	plt.xticks(ypos, Date)
	plt.yticks(TeaMade)
	plt.xlabel('Days',color="#E2A089")
	plt.ylabel('Tea Made',color="#E2A089")
	# Show Plot
	plt.show()
	#df=pd.read_csv("data1.csv")



####################################################################################							Monthly Graph

def Monthly_Graph():
	'''x=[1,2,3]
				y=[1,4,9]
				plt.plot(x,y)
				plt.show()'''
	# Read CSV into pandas
	#data = pd.read_csv(r"officedata.csv")
	data=pd.read_excel(r"officedata.xlsx")
	data.head()
	df = pd.DataFrame(data)
	##print(df)
	#df['ConvertedDate']=df['Date'].astype(str)
	Date= df['Date'].dt.strftime('%b-%d').head(31)
	#print(Date)
	TeaPlucked = df['Tea Plucked'].head(31)
	#print(TeaPlucked)
	TeaMade=df['Tea Made'].head(31)
	#print(TeaMade)
	# Figure Size
	fig = plt.figure(figsize =(30, 30))
	# Horizontal Bar Plot
	ypos=np.arange(len(Date))
	#print("length of ypos is: ",ypos)
	plt.bar(ypos, TeaMade,color="#A934BD")
	plt.xticks(ypos, Date)	


	# Show Plot
	plt.show()
def Yearly_Graph():
	'''x=[1,2,3]
				y=[1,4,9]
				plt.plot(x,y)
				plt.show()'''
	# Read CSV into pandas
	#data = pd.read_csv(r"officedata.csv")
	data=pd.read_excel(r"officedata.xlsx")
	data.head()
	df = pd.DataFrame(data)
	##print(df)
	#df['ConvertedDate']=df['Date'].astype(str)
	#Date=df.groupby(df['Date'].dt.strftime('%B'))['Tea Made'].sum()
	'''df['month'] = pd.to_datetime(df['date']).dt.to_period('M')
	# Group by the month and plot
	df.groupby('month')['model'].count().plot.bar();'''
	Date=df['Date'].dt.strftime('%B') 
	#print("This is the date")
	#print(Date)
	TeaPlucked = df['Tea Plucked'].head(365)
	##print(TeaPlucked)
	TeaMade= df.groupby(df['Date'].dt.strftime('%B'))['Tea Made'].plot.bar() #sort_values() to sort values
	##print(TeaMade)
	# Figure Size
	#fig = plt.figure(figsize =(30, 30))
	# Horizontal Bar Plot
	#ypos=np.arange(len(Date))
	##print("length of ypos is: ",ypos)
	#plt.bar(df[Date], TeaMade)
	#plt.xticks(ypos,Date)	


	# Show Plot
	plt.show()

def Weekly_progress():
	'''x=[1,2,3]
				y=[1,4,9]
				plt.plot(x,y)
				plt.show()'''
	# Read CSV into pandas
	#data = pd.read_csv(r"officedata.csv")
	data=pd.read_excel(r"officedata.xlsx")
	data.head()
	df = pd.DataFrame(data)
	##print(df)
	#df['ConvertedDate']=df['Date'].astype(str)
	Date= df['Date'].dt.strftime('%d-%b-%Y').head(7)

	#print(Date)
	TeaPlucked = df['Tea Plucked'].head(7)
	#print(TeaPlucked)
	TeaMade=df['Tea Made'].head(7)
	#print(TeaMade)
	# Figure Size
	fig = plt.figure(figsize =(7,7))
	# Horizontal Bar Plot
	ypos=np.arange(len(Date))
	#print("length of ypos is: ",ypos)
	for i in range(len(ypos)):
		plt.text(i,TeaMade[i],TeaMade[i])
	#plt.bar(ypos, TeaMade, color='#6C5B87',width=0.5)
	#plt.bar(ypos, TeaMade, color='#C6808B',width=0.5)
	
	plt.plot(ypos, TeaMade, color='#6C5B87')
	
	plt.xticks(ypos, Date)
	plt.yticks(TeaMade)
	plt.xlabel('Days',color="#E2A089")
	plt.ylabel('Tea Made',color="#E2A089")
	plt.legend()
	# Show Plot
	plt.show()
	#df=pd.read_csv("data1.csv")


apps_frame= LabelFrame(root,bd=0)
apps_frame.pack(side="bottom")
apps_frame.configure(bg="#1D3F49")

Weekly_Graph_button=Button(apps_frame, text="Weekly Analysis", command=Weekly_Graph).pack(side=LEFT)
Monthly_Graph_button=Button(apps_frame, text="Montly Analysis", command=Monthly_Graph).pack(side=LEFT)
Yearly_Graph_button=Button(apps_frame, text="Yearly Analysis", command=Yearly_Graph).pack(side=LEFT)
Weekly_progress_Graph_button=Button(apps_frame, text="Weekly progress", command=Weekly_progress).pack(side=LEFT)







error_label=Label(root,text="")
error_label.pack(pady=20)













######################################################################################## testing
testdir="C:\\Users\\"+os.getlogin()+"\\Documents"+"\\TEST\\officedata.xlsx"
df=pd.read_excel(testdir)	#for test 
create_tree_view_for_excel_sheet()
######################################################################################## testing end







root.mainloop()